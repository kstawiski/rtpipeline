#!/umed-projekty/rtpipeline/.venv/bin/python
from __future__ import annotations

import argparse
import json
import math
import time
from collections import Counter
from pathlib import Path

import numpy as np
import pandas as pd
import pyarrow.parquet as pq
from openpyxl import load_workbook
from scipy.stats import spearmanr


PROSTATE_RADIOMICS_CANDIDATES = [
    Path("/home/kgs24/Prostate_Radiomics_Kopernik_ASTRO2026/results/radiomics_ts_all.parquet"),
]
PROSTATE_DVH_CANDIDATES = [
    Path("/projekty/Prostata/Data_Snakemake/dvh_summary.xlsx"),
]
ICC_RESULTS_CANDIDATES = [
    Path("/umed-projekty/rtpipeline/manuscript/analysis/data/icc_results.parquet"),
    Path("/home/kgs24/rtpipeline_manuscript/analysis/data/icc_results.parquet"),
]

ROOT = Path("/umed-projekty/rtpipeline")
TABLES_DIR = ROOT / "manuscript" / "analysis" / "tables"
DATA_DIR = ROOT / "manuscript" / "analysis" / "data"

OUT_COVERAGE = TABLES_DIR / "45_dose_radiomic_response_prostata_roi_coverage.csv"
OUT_ROI_FEATURES = TABLES_DIR / "45_dose_radiomic_response_prostata_roi_feature_screen.csv"
OUT_FEATURES = TABLES_DIR / "45_dose_radiomic_response_prostata_feature_screen.csv"
OUT_CLASS = TABLES_DIR / "45_dose_radiomic_response_prostata_class_summary.csv"
OUT_FAMILY = TABLES_DIR / "45_dose_radiomic_response_prostata_family_summary.csv"
OUT_MANIFEST = DATA_DIR / "45_dose_radiomic_response_prostata_manifest.json"


def pick_existing(candidates: list[Path]) -> Path:
    existing = [path for path in candidates if path.exists()]
    if not existing:
        raise FileNotFoundError(f"None of the candidate paths exist: {candidates}")
    if len(existing) == 1:
        return existing[0]
    return max(existing, key=lambda path: path.stat().st_mtime)


def normalize_text(value: object) -> str | None:
    if pd.isna(value):
        return None
    text = str(value).strip().lower()
    out = "".join(ch if ch.isalnum() else "_" for ch in text).strip("_")
    return out or None


def normalize_patient_id(value: object) -> str | None:
    if pd.isna(value):
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return str(int(float(text)))
    except Exception:
        return text


def infer_family(feature_name: str) -> str:
    lowered = feature_name.lower()
    for family in ("shape", "firstorder", "glcm", "glrlm", "glszm", "gldm", "ngtdm"):
        if f"_{family}_" in lowered or lowered.startswith(f"{family}_"):
            return family
    return "other"


def bh_fdr(p_values: pd.Series) -> pd.Series:
    arr = pd.to_numeric(p_values, errors="coerce").to_numpy(dtype=float)
    out = np.full(arr.shape, np.nan, dtype=float)
    valid = np.isfinite(arr)
    if not valid.any():
        return pd.Series(out, index=p_values.index)
    p = arr[valid]
    order = np.argsort(p)
    ranked = p[order]
    n = len(ranked)
    q = np.empty(n, dtype=float)
    prev = 1.0
    for i in range(n - 1, -1, -1):
        rank = i + 1
        val = min(prev, ranked[i] * n / rank)
        q[i] = val
        prev = val
    adjusted = np.empty(n, dtype=float)
    adjusted[order] = q
    out[np.where(valid)[0]] = adjusted
    return pd.Series(out, index=p_values.index)


def load_prostate_radiomics() -> tuple[Path, pd.DataFrame, list[str]]:
    path = pick_existing(PROSTATE_RADIOMICS_CANDIDATES)
    print(f"[45] Loading radiomics schema: {path}", flush=True)
    parquet = pq.ParquetFile(path)
    schema = parquet.schema.names
    feature_cols = [col for col in schema if col.startswith("original_")]
    read_cols = ["patient_id", "roi_name", *feature_cols]
    print(f"[45] Reading radiomics table with {len(feature_cols)} original features", flush=True)
    table = parquet.read(columns=read_cols, use_threads=False)
    df = table.to_pandas()
    df["patient_id_norm"] = df["patient_id"].map(normalize_patient_id)
    df["roi_norm"] = df["roi_name"].map(normalize_text)
    df = df.dropna(subset=["patient_id_norm", "roi_norm"])
    df = df.drop_duplicates(["patient_id_norm", "roi_norm"], keep="first")
    return path, df, feature_cols


def load_prostate_dvh(dose_metric: str) -> tuple[Path, pd.DataFrame]:
    path = pick_existing(PROSTATE_DVH_CANDIDATES)
    print(f"[45] Loading DVH workbook: {path}", flush=True)
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    required = {"patient", "ROI_Name", dose_metric, "DmaxGy", "D95Gy", "Volume (cm³)"}
    missing = sorted(required - set(header))
    if missing:
        raise RuntimeError(f"DVH workbook missing columns: {missing}")
    idx = {name: header.index(name) for name in required}
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rows.append(
            {
                "patient_id_norm": normalize_patient_id(row[idx["patient"]]),
                "roi_name_dvh": row[idx["ROI_Name"]],
                "roi_norm": normalize_text(row[idx["ROI_Name"]]),
                dose_metric: row[idx[dose_metric]],
                "DmaxGy": row[idx["DmaxGy"]],
                "D95Gy": row[idx["D95Gy"]],
                "Volume_cm3": row[idx["Volume (cm³)"]],
            }
        )
    df = pd.DataFrame(rows)
    df = df.dropna(subset=["patient_id_norm", "roi_norm", dose_metric])
    df = df.drop_duplicates(["patient_id_norm", "roi_norm"], keep="first")
    return path, df


def load_icc_summary(matched_rois: list[str], feature_cols: list[str]) -> tuple[Path, pd.DataFrame]:
    path = pick_existing(ICC_RESULTS_CANDIDATES)
    print(f"[45] Loading ICC summary: {path}", flush=True)
    icc = pd.read_parquet(path)
    icc["cohort_norm"] = icc["cohort"].map(normalize_text)
    icc = icc.loc[icc["cohort_norm"] == "prostata"].copy()
    roi_col = "roi_name" if "roi_name" in icc.columns else "roi"
    icc["roi_norm"] = icc[roi_col].map(normalize_text)
    icc = icc.loc[icc["roi_norm"].isin(matched_rois)].copy()
    if "image_type" in icc.columns:
        ct_mask = icc["image_type"].astype(str).str.upper().eq("CT")
        if ct_mask.any():
            icc = icc.loc[ct_mask].copy()
    icc = icc.loc[icc["feature_name"].isin(feature_cols)].copy()
    class_col = "classification" if "classification" in icc.columns else "robustness_class"
    icc["robustness_class"] = icc[class_col].astype(str).str.title()
    if "feature_family" in icc.columns:
        icc["feature_family"] = icc["feature_family"].astype(str).str.lower()
    else:
        icc["feature_family"] = icc["feature_name"].map(infer_family)
    agg_rows = []
    for feature_name, group in icc.groupby("feature_name", sort=False):
        agg_rows.append(
            {
                "feature_name": feature_name,
                "feature_family": Counter(group["feature_family"]).most_common(1)[0][0],
                "robustness_class": Counter(group["robustness_class"]).most_common(1)[0][0],
                "icc_roi_count": int(group["roi_norm"].nunique()),
                "icc_median": float(pd.to_numeric(group["icc"], errors="coerce").median()),
            }
        )
    return path, pd.DataFrame(agg_rows)


def build_matched_table(dose_metric: str) -> tuple[dict[str, str], pd.DataFrame, list[str]]:
    rad_path, rad, feature_cols = load_prostate_radiomics()
    dvh_path, dvh = load_prostate_dvh(dose_metric)
    matched = rad.merge(
        dvh,
        on=["patient_id_norm", "roi_norm"],
        how="inner",
        validate="one_to_one",
    )
    sources = {
        "prostate_radiomics": str(rad_path),
        "prostate_dvh": str(dvh_path),
    }
    return sources, matched, feature_cols


def summarize_coverage(matched: pd.DataFrame, dose_metric: str) -> pd.DataFrame:
    rows = []
    for roi_norm, group in matched.groupby("roi_norm", sort=True):
        dose = pd.to_numeric(group[dose_metric], errors="coerce")
        rows.append(
            {
                "roi_norm": roi_norm,
                "roi_name_example": group["roi_name"].dropna().astype(str).mode().iloc[0],
                "roi_name_dvh_example": group["roi_name_dvh"].dropna().astype(str).mode().iloc[0],
                "n_rows": int(len(group)),
                "n_patients": int(group["patient_id_norm"].nunique()),
                "dose_min": float(dose.min()),
                "dose_median": float(dose.median()),
                "dose_max": float(dose.max()),
                "dose_iqr": float(dose.quantile(0.75) - dose.quantile(0.25)),
            }
        )
    out = pd.DataFrame(rows).sort_values(["n_patients", "roi_norm"], ascending=[False, True])
    return out


def run_roi_feature_screen(
    matched: pd.DataFrame,
    feature_cols: list[str],
    coverage: pd.DataFrame,
    dose_metric: str,
    min_roi_n: int,
) -> pd.DataFrame:
    keep_rois = coverage.loc[coverage["n_patients"] >= min_roi_n, "roi_norm"].tolist()
    work = matched.loc[matched["roi_norm"].isin(keep_rois)].copy()
    blocks = []
    for roi_norm, roi_df in work.groupby("roi_norm", sort=True):
        rows = []
        dose = pd.to_numeric(roi_df[dose_metric], errors="coerce")
        for feature_name in feature_cols:
            feature = pd.to_numeric(roi_df[feature_name], errors="coerce")
            sub = pd.DataFrame({"dose": dose, "feature": feature}).dropna()
            if len(sub) < max(min_roi_n, 8):
                continue
            if sub["dose"].nunique() < 2 or sub["feature"].nunique() < 2:
                continue
            rho, p_value = spearmanr(sub["dose"], sub["feature"], nan_policy="omit")
            if not np.isfinite(rho) or not np.isfinite(p_value):
                continue
            rows.append(
                {
                    "roi_norm": roi_norm,
                    "feature_name": feature_name,
                    "feature_family": infer_family(feature_name),
                    "n_obs": int(len(sub)),
                    "n_patients": int(roi_df["patient_id_norm"].nunique()),
                    "spearman_rho": float(rho),
                    "abs_spearman_rho": float(abs(rho)),
                    "p_value": float(p_value),
                }
            )
        block = pd.DataFrame(rows)
        block["fdr_bh_within_roi"] = bh_fdr(block["p_value"])
        blocks.append(block)
    out = pd.concat(blocks, ignore_index=True)
    out = out.sort_values(["roi_norm", "abs_spearman_rho", "p_value"], ascending=[True, False, True])
    return out


def aggregate_feature_summary(roi_feature_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for feature_name, group in roi_feature_df.groupby("feature_name", sort=False):
        median_rho = float(group["spearman_rho"].median())
        pos = int((group["spearman_rho"] > 0).sum())
        neg = int((group["spearman_rho"] < 0).sum())
        best_row = group.sort_values(["abs_spearman_rho", "p_value"], ascending=[False, True]).iloc[0]
        rows.append(
            {
                "feature_name": feature_name,
                "feature_family": group["feature_family"].iloc[0],
                "n_rois_tested": int(group["roi_norm"].nunique()),
                "median_spearman_rho": median_rho,
                "median_abs_spearman_rho": float(group["abs_spearman_rho"].median()),
                "max_abs_spearman_rho": float(group["abs_spearman_rho"].max()),
                "n_rois_nominal_p_lt_0_05": int((group["p_value"] < 0.05).sum()),
                "n_rois_fdr_lt_0_10": int((group["fdr_bh_within_roi"] < 0.10).sum()),
                "sign_consistency": float(max(pos, neg) / len(group)),
                "best_roi": best_row["roi_norm"],
                "best_roi_rho": float(best_row["spearman_rho"]),
                "best_roi_p_value": float(best_row["p_value"]),
            }
        )
    return pd.DataFrame(rows).sort_values(
        ["median_abs_spearman_rho", "n_rois_fdr_lt_0_10", "max_abs_spearman_rho"],
        ascending=[False, False, False],
    )


def summarize_feature_groups(feature_df: pd.DataFrame, group_col: str) -> pd.DataFrame:
    rows = []
    for key, group in feature_df.groupby(group_col, dropna=False, sort=True):
        rows.append(
            {
                group_col: key,
                "n_features": int(len(group)),
                "median_abs_spearman_rho": float(group["median_abs_spearman_rho"].median()),
                "q75_abs_spearman_rho": float(group["median_abs_spearman_rho"].quantile(0.75)),
                "n_features_with_any_roi_fdr_lt_0_10": int((group["n_rois_fdr_lt_0_10"] > 0).sum()),
                "n_features_with_any_roi_nominal_p_lt_0_05": int((group["n_rois_nominal_p_lt_0_05"] > 0).sum()),
            }
        )
    return pd.DataFrame(rows).sort_values("median_abs_spearman_rho", ascending=False)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Exploratory Prostata dose-radiomic response screen using matched DVH and radiomics."
    )
    parser.add_argument("--dose-metric", default="DmeanGy")
    parser.add_argument("--min-roi-n", type=int, default=28)
    args = parser.parse_args()

    start = time.time()
    TABLES_DIR.mkdir(parents=True, exist_ok=True)
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    print("[45] Building matched Prostata dose-radiomics table", flush=True)
    sources, matched, feature_cols = build_matched_table(args.dose_metric)
    print(
        f"[45] Matched rows={len(matched)} patients={matched['patient_id_norm'].nunique()} rois={matched['roi_norm'].nunique()}",
        flush=True,
    )
    coverage = summarize_coverage(matched, args.dose_metric)
    print("[45] ROI coverage summary ready", flush=True)
    icc_path, icc_summary = load_icc_summary(coverage["roi_norm"].tolist(), feature_cols)
    print(f"[45] ICC summary rows={len(icc_summary)}", flush=True)
    roi_feature_df = run_roi_feature_screen(
        matched=matched,
        feature_cols=feature_cols,
        coverage=coverage,
        dose_metric=args.dose_metric,
        min_roi_n=args.min_roi_n,
    )
    print(f"[45] ROI-level feature screen rows={len(roi_feature_df)}", flush=True)
    feature_df = aggregate_feature_summary(roi_feature_df)
    print(f"[45] Aggregated feature screen rows={len(feature_df)}", flush=True)
    feature_df = feature_df.merge(icc_summary, on=["feature_name"], how="left", suffixes=("", "_icc"))
    feature_df["feature_family"] = feature_df["feature_family"].fillna(feature_df["feature_family_icc"])
    feature_df = feature_df.drop(columns=["feature_family_icc"], errors="ignore")
    feature_df["robustness_class"] = feature_df["robustness_class"].fillna("Unknown")

    class_summary = summarize_feature_groups(feature_df, "robustness_class")
    family_summary = summarize_feature_groups(feature_df, "feature_family")

    coverage.to_csv(OUT_COVERAGE, index=False)
    roi_feature_df.to_csv(OUT_ROI_FEATURES, index=False)
    feature_df.to_csv(OUT_FEATURES, index=False)
    class_summary.to_csv(OUT_CLASS, index=False)
    family_summary.to_csv(OUT_FAMILY, index=False)

    manifest = {
        "analysis": "#45 dose-radiomic response",
        "cohort": "Prostata",
        "dose_metric": args.dose_metric,
        "min_roi_n": args.min_roi_n,
        "runtime_seconds": round(time.time() - start, 2),
        "sources": {**sources, "icc_results": str(icc_path)},
        "matched_rows": int(len(matched)),
        "matched_patients": int(matched["patient_id_norm"].nunique()),
        "matched_rois": int(matched["roi_norm"].nunique()),
        "roi_feature_path": str(OUT_ROI_FEATURES),
        "feature_count": int(len(feature_df)),
        "coverage_path": str(OUT_COVERAGE),
        "feature_path": str(OUT_FEATURES),
        "class_summary_path": str(OUT_CLASS),
        "family_summary_path": str(OUT_FAMILY),
        "top_feature_by_abs_rho": (
            None
            if feature_df.empty
            else feature_df.iloc[0][
                [
                    "feature_name",
                    "feature_family",
                    "median_spearman_rho",
                    "median_abs_spearman_rho",
                    "n_rois_fdr_lt_0_10",
                    "robustness_class",
                ]
            ].to_dict()
        ),
    }
    OUT_MANIFEST.write_text(json.dumps(manifest, indent=2) + "\n")

    print(json.dumps(manifest, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
